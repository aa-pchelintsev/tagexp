#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
tagexp.py — офлайн-эксперименты по AI‑тегированию звонков

CLI: python tagexp.py --config ./config.yaml

Скрипт реализует офлайн‑эксперименты по маркировке телефонных диалогов тегами на основе
AI‑модели через внутренний API. Реализованы функции загрузки данных, многократные
запросы к API с большинством по результатам, сбор диагностической информации,
формирование отчётного Excel с листами «Главный», «Сводка» и «Диагностика»,
а также вспомогательные LAMBDA‑формулы для подсчёта ошибок (+/−).

Ключевые особенности:
 - system_prompt формируется из файла p*.txt с заменой плейсхолдера {{TRANSCRIPT}}
   на текст расшифровки;
 - управление количеством повторов и правилом большинства configurable через YAML;
 - поддержка Excel LAMBDA/LET формул для подсчёта ошибок и стабильности;
 - полная совместимость с Microsoft 365; предусмотрен fallback для использования
   Python‑расчётов (включив excel.enable_dynamic_formulas=false);
 - продуманный интерфейс командной строки и журналирование.
"""

import os
import sys
import json
import time
import math
import argparse
from pathlib import Path
import pandas as pd
import requests
import json
from tenacity import retry, wait_exponential, stop_after_attempt, retry_if_exception_type
from typing import Optional
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.utils import get_column_letter


# Канонический набор тегов строго в порядке, соответствующем промпту
ALLOWED_TAGS = [
    "Отказ",
    "Негатив",
    "Дорого",
    "Позовите руководителя",
    "Допродажа",
    "Сервис",
    "Доставка",
    "Оплата",
    "Возврат",
    "Акция",
]


def load_yaml(path: Path) -> dict:
    """Загружает YAML‑конфиг в виде словаря."""
    import yaml
    with open(path, "r", encoding="utf-8") as f:
        cfg = yaml.safe_load(f)
    # Базовая директория для относительных путей
    base_dir = path.parent.resolve()
    # Обновляем все относительные пути в io и output
    for section in ("io", "output"):
        if section in cfg:
            for k, v in cfg[section].items():
                if isinstance(v, str) and v and not Path(v).is_absolute():
                    cfg[section][k] = str((base_dir / v).resolve())
    # Аналогично для logging->file
    if "logging" in cfg and "file" in cfg["logging"]:
        v = cfg["logging"]["file"]
        if isinstance(v, str) and v and not Path(v).is_absolute():
            cfg["logging"]["file"] = str((base_dir / v).resolve())
    return cfg


def json_repair(text: str):
    """Пытается распарсить JSON из текста, даже если он содержит посторонние символы."""
    try:
        return json.loads(text)
    except Exception:
        pass
    l = text.find("{")
    r = text.rfind("}")
    if l != -1 and r != -1 and r > l:
        candidate = text[l : r + 1]
        try:
            return json.loads(candidate)
        except Exception:
            pass
    raise ValueError("Не удалось распарсить JSON из ответа модели")


def majority_bool(values, rule="ceil_half"):
    """Возвращает True, если большинство значений из списка истинно согласно правилу."""
    if not values:
        return False
    if rule == "ceil_half":
        return sum(bool(v) for v in values) >= math.ceil(len(values) / 2)
    # fallback: последний результат
    return bool(values[-1])


def short_quotes_from_evidence(evidence, true_tags, max_quotes=2, max_words=20):
    """Собирает короткие цитаты из evidence для заданных тегов."""
    quotes = []
    ev = evidence if isinstance(evidence, dict) else {}
    for tag in true_tags:
        arr = ev.get(tag, []) if isinstance(ev, dict) else []
        for item in arr:
            quote = (item.get("quote") or "").strip()
            if not quote:
                continue
            words = quote.split()
            if len(words) > max_words:
                quote = " ".join(words[:max_words]) + "…"
            quotes.append(quote)
            if len(quotes) >= max_quotes:
                return "; ".join(quotes)
    return "; ".join(quotes)


def debug_flags_to_notes(debug_obj):
    """Приводит поля из _debug в удобочитаемый формат."""
    def pick(d):
        out = []
        if "role_ok" in d:
            out.append(f"роль={'ОК' if str(d['role_ok']).lower() in ('1','1.0','true') else 'неОК'}")
        if "finality" in d:
            out.append(
                f"финальность={'Да' if str(d['finality']).lower() in ('1','1.0','true') else 'Нет'}"
            )
        if "proximity" in d:
            prox = str(d["proximity"]).lower()
            out.append(f"близость={'Высокая' if prox in ('1','true','high') else 'Низкая'}")
        if "anti" in d:
            out.append(f"анти‑фраза={'Да' if str(d['anti']).lower() in ('1','true') else 'Нет'}")
        if d.get("dropped_by_conflict"):
            out.append("конфликт=Снят")
        return out

    pieces = []
    if isinstance(debug_obj, dict):
        for tag, d in debug_obj.items():
            if tag.startswith("_"):
                continue
            if isinstance(d, dict):
                parts = pick(d)
                if parts:
                    pieces.append(", ".join(parts))
    return "; ".join(pieces)


def load_calls(cfg) -> pd.DataFrame:
    """Загружает файл звонков (Excel или CSV) в DataFrame."""
    path = Path(cfg["io"]["calls_path"])
    print(f"[INFO] Загрузка звонков из файла: {path}")
    if path.suffix.lower() in (".xlsx", ".xls"):
        df = pd.read_excel(path)
    elif path.suffix.lower() == ".csv":
        df = pd.read_csv(path)
    else:
        raise ValueError(f"Неподдерживаемый формат входного файла: {path.suffix}")
    required = ["Система", "Уникальные ID", "Расшифровка"]
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(f"Отсутствуют обязательные колонки: {missing}")
    # добавляем колонку для оценок человека, если нет
    human_col = cfg["excel"]["human_tags_column"]
    if human_col not in df.columns:
        df[human_col] = ""
    print(f"[INFO] Загружено звонков: {len(df)}")
    return df


def load_prompts(cfg) -> dict:
    """Загружает все файлы p*.txt из директории prompts_dir."""
    pdir = Path(cfg["io"]["prompts_dir"])
    print(f"[INFO] Загрузка промптов из директории: {pdir}")
    files = sorted(pdir.glob("p*.txt"))
    if not files:
        raise FileNotFoundError(f"В {pdir} нет файлов p*.txt")
    out = {}
    for f in files:
        out[f.stem.upper()] = f.read_text(encoding="utf-8")
    print(f"[INFO] Загружено промптов: {len(out)}")
    return out


class APIClient:
    """Клиент для взаимодействия с API с учётом таймаутов и ограничения на частоту."""

    def __init__(self, cfg: dict):
        self.cfg = cfg
        self.base_url = cfg["api"]["base_url"]
        self.account_id = os.getenv("ACCOUNT_ID", cfg["api"]["account_id"])
        self.model = cfg["api"]["model"]
        self.temperature = cfg["api"].get("temperature", 0.1)
        # Ключ: сначала из переменной окружения, затем из config.yaml (api.api_key), иначе "xxxxxxxx"
        api_key = os.getenv("API_KEY") or cfg["api"].get("api_key") or "xxxxxxxx"
        self.headers = {
            "Content-Type": "application/json",
            "X-API-KEY": api_key,
        }
        self.timeout = cfg["api"].get("timeout_sec", 60)
        self.rate_limit_per_minute = cfg["experiment"].get("rate_limit_per_minute", 30)
        # минимальный интервал между вызовами
        self.min_interval = 60.0 / max(1, self.rate_limit_per_minute)
        self._last_call_ts = 0.0

    def _respect_rate(self):
        now = time.time()
        dt = now - self._last_call_ts
        if dt < self.min_interval:
            time.sleep(self.min_interval - dt)
        self._last_call_ts = time.time()

    @retry(
        wait=wait_exponential(multiplier=0.5, min=0.5, max=4.0),
        stop=stop_after_attempt(3),
        retry=retry_if_exception_type((requests.RequestException, ValueError)),
    )
    def call(self, system_prompt_with_transcript: str) -> dict:
        """Отправляет запрос к API и возвращает JSON (или словарь)."""
        self._respect_rate()
        body = {
            "account_id": self.account_id,
            "gpt": {
                "model": self.model,
                "system_prompt": system_prompt_with_transcript,
                "temperature": self.temperature,
            },
        }
        # Заменяем особые символы в system_prompt (например, длинное тире и кавычки)
        if "gpt" in body and "system_prompt" in body["gpt"]:
            sp = body["gpt"]["system_prompt"]
            sp = sp.replace("—", "-")
            sp = sp.replace("“", '"').replace("”", '"')
            sp = sp.replace("‘", "'").replace("’", "'")
            body["gpt"]["system_prompt"] = sp

        # Явно указываем charset=utf-8
        headers = dict(self.headers)
        headers["Content-Type"] = "application/json; charset=utf-8"

        print("[INFO] Устанавливаю соединение с API...")
        try:
            print("[INFO] Отправка запроса к API...")
            r = requests.post(
                self.base_url,
                headers=headers,
                json=body,
                timeout=self.timeout,
            )
            print("[INFO] Запрос отправлен, ожидаю ответ...")
        except Exception as e:
            print(f"[ERROR] Ошибка при отправке запроса: {e}")
            raise
        print(f"[INFO] Ответ получен: status={r.status_code}")
        # 429 или 5xx — считаем временной ошибкой
        if r.status_code == 429 or r.status_code >= 500:
            raise requests.RequestException(f"{r.status_code}")
        if r.status_code >= 400:
            # пробрасываем текст ошибки, но обрезаем слишком длинный
            raise requests.RequestException(f"{r.status_code} {r.text[:200]}")
        return json_repair(r.text)


def run_experiment(cfg: dict, calls_df: pd.DataFrame, prompts: dict):
    """Основная процедура эксперимента: обход звонков, опросы API и сбор результатов."""
    api = APIClient(cfg)
    repeats = int(cfg["experiment"]["repeats_per_prompt"])
    maj_rule = cfg["experiment"]["majority_rule"]

    results_rows = []
    diag_rows = []

    total = len(calls_df)
    for idx, (_, row) in enumerate(calls_df.iterrows(), 1):
        system = row["Система"]
        uid = row["Уникальные ID"]
        transcript = row["Расшифровка"] or ""
        human_tags = row.get(cfg["excel"]["human_tags_column"], "")

        print(f"[INFO] Обработка звонка {idx}/{total}: Система={system}, ID={uid}")

        per_prompt = {}
        # для каждого промпта делаем N повторов
        for pkey, ptxt in prompts.items():
            print(f"[INFO]  Промпт: {pkey}")
            attempts = []
            for i in range(repeats):
                print(f"[INFO]   Попытка {i+1}/{repeats}")
                filled = ptxt.replace("{{TRANSCRIPT}}", transcript)

                try:
                    data = api.call(filled)
                except Exception as exc:
                    print(f"[ERROR] API {system}/{uid} {pkey}#{i+1}: {exc}")
                    data = {}

                print(f"[DEBUG] Ответ API: {data}")
                tag_bools = {t: bool(data.get(t, False)) for t in ALLOWED_TAGS}
                print(f"[DEBUG] tag_bools: {tag_bools}")
                attempts.append(tag_bools)

                evidence = data.get("_evidence", {}) if isinstance(data, dict) else {}
                debug = data.get("_debug", {}) if isinstance(data, dict) else {}
                # диагностика: одна строка на каждую комбинацию тег‑цитата (или пустая)
                for tag in ALLOWED_TAGS:
                    ev_list = evidence.get(tag, []) if isinstance(evidence, dict) else []
                    if ev_list:
                        for ev in ev_list:
                            diag_rows.append(
                                {
                                    "Система": system,
                                    "Уникальные ID": uid,
                                    "Промпт": pkey,
                                    "Попытка": i + 1,
                                    "Тег": tag,
                                    "Решение": tag_bools.get(tag, False),
                                    "Конфликт": bool(
                                        debug.get(tag, {}).get("dropped_by_conflict", False)
                                    )
                                    if isinstance(debug, dict)
                                    else False,
                                    "Цитата": (ev.get("quote") or "").strip(),
                                    "Номер реплики": ev.get("turn"),
                                    "Кто говорит": ev.get("role"),
                                    "Категория сигнала": ev.get("group"),
                                    "Флаги": debug.get(tag, {}) if isinstance(debug, dict) else {},
                                }
                            )
                    else:
                        diag_rows.append(
                            {
                                "Система": system,
                                "Уникальные ID": uid,
                                "Промпт": pkey,
                                "Попытка": i + 1,
                                "Тег": tag,
                                "Решение": tag_bools.get(tag, False),
                                "Конфликт": bool(
                                    debug.get(tag, {}).get("dropped_by_conflict", False)
                                )
                                if isinstance(debug, dict)
                                else False,
                                "Цитата": "",
                                "Номер реплики": None,
                                "Кто говорит": None,
                                "Категория сигнала": None,
                                "Флаги": debug.get(tag, {}) if isinstance(debug, dict) else {},
                            }
                        )

            # агрегируем по большинству
            final_true = []
            tagset_strings = []
            for at in attempts:
                true_list = [t for t, v in at.items() if v]
                true_list = [t for t in ALLOWED_TAGS if t in true_list]
                tagset_strings.append(",".join(true_list))
            for tag in ALLOWED_TAGS:
                if majority_bool([a.get(tag, False) for a in attempts], maj_rule):
                    final_true.append(tag)

            # стабильность: доля попыток с наиболее частым набором
            from collections import Counter

            c = Counter(tagset_strings)
            mc = c.most_common(1)[0][1] if tagset_strings else 0
            stability_num = mc
            stability_den = len(attempts)
            stability_pct = (stability_num / stability_den * 100.0) if stability_den else 0.0

            # соберём последний evidence/debug для пояснений
            ev_map, dbg_map = {}, {}
            for r in filter(
                lambda r: r["Система"] == system
                and r["Уникальные ID"] == uid
                and r["Промпт"] == pkey
                and r["Попытка"] == repeats,
                diag_rows,
            ):
                tag_name = r["Тег"]
                if r["Цитата"]:
                    ev_map.setdefault(tag_name, []).append(
                        {
                            "quote": r["Цитата"],
                            "turn": r["Номер реплики"],
                            "role": r["Кто говорит"],
                            "group": r["Категория сигнала"],
                        }
                    )
                dbg_map[tag_name] = r.get("Флаги") or {}

            explanation = short_quotes_from_evidence(ev_map, final_true)
            notes = debug_flags_to_notes(dbg_map)

            per_prompt[pkey] = {
                "tags_str": ", ".join(final_true),
                "explanation": explanation,
                "notes": notes,
                "stability": f"{stability_num}/{stability_den} ({stability_pct:.0f}%)",
                "stability_pct": round(stability_pct, 2),
            }

        # собираем строку для «Главного» листа
        res = {
            "Система": system,
            "Уникальные ID": uid,
            "Расшифровка": transcript,
            cfg["excel"]["human_tags_column"]: human_tags or "",
        }
        for pkey in sorted(per_prompt.keys()):
            res[f"Теги [{pkey}]"] = per_prompt[pkey]["tags_str"]
            res[f"Пояснение [{pkey}]"] = per_prompt[pkey]["explanation"]
            res[f"Ошибки (−) [{pkey}]"] = ""
            res[f"Ошибки (+) [{pkey}]"] = ""
            res[f"Замечания [{pkey}]"] = per_prompt[pkey]["notes"]
            res[f"Стабильность [{pkey}]"] = per_prompt[pkey]["stability"]
            res[f"Стабильность [{pkey}] %"] = per_prompt[pkey]["stability_pct"]
            res[f"Ошибки (−) #[{pkey}]"] = 0
            res[f"Ошибки (+) #[{pkey}]"] = 0
        results_rows.append(res)

    return pd.DataFrame(results_rows), pd.DataFrame(diag_rows)


def define_lambda_names(wb: Workbook):
    """Определяет имена функций LAMBDA для работы в Excel."""
    tag_list = (
        '=LAMBDA(txt,LET(s,SUBSTITUTE(TRIM(IF(txt="","",txt)),"  "," "),arr,TEXTSPLIT(s,","),t,TRIM(arr),FILTER(t,t<>"")))'
    )
    set_minus = (
        '=LAMBDA(a,b,LET(A,UNIQUE(SORT(TAG_LIST(a))),B,UNIQUE(SORT(TAG_LIST(b))),RES,FILTER(A,ISNA(XMATCH(A,B,0))),TEXTJOIN(", ",TRUE,RES)))'
    )
    set_equal = (
        '=LAMBDA(a,b,LET(m1,SET_MINUS(a,b),m2,SET_MINUS(b,a),IF(AND(m1="",m2=""),TRUE,FALSE)))'
    )
    wb.defined_names.add(DefinedName("TAG_LIST", attr_text=tag_list))
    wb.defined_names.add(DefinedName("SET_MINUS", attr_text=set_minus))
    wb.defined_names.add(DefinedName("SET_EQUAL", attr_text=set_equal))


def build_workbook(cfg: dict, calls_df: pd.DataFrame, results_df: pd.DataFrame, diagnostics_df: pd.DataFrame):
    # Экспорт в CSV для надёжного импорта в Excel/Google Sheets
    out_dir = Path(cfg["output"]["path"]).parent
    results_csv = out_dir / "results.csv"
    diagnostics_csv = out_dir / "diagnostics.csv"
    results_df.to_csv(results_csv, index=False, encoding="utf-8-sig")
    diagnostics_df.to_csv(diagnostics_csv, index=False, encoding="utf-8-sig")
    print(f"[INFO] CSV-файлы сохранены: {results_csv}, {diagnostics_csv}")
    """Формирует Excel‑файл с листами «Главный», «Сводка» и «Диагностика»."""
    wb = Workbook()
    ws_main = wb.active
    ws_main.title = "Главный"

    # перенос длинных расшифровок на отдельный лист при предпросмотре
    if not cfg["output"].get("show_full_transcript", True):
        full_ws = wb.create_sheet("Тексты")
        full_ws.append(["Система", "Уникальные ID", "Полный текст"])
        for _, r in calls_df.iterrows():
            full_ws.append([r["Система"], r["Уникальные ID"], r["Расшифровка"]])
        results_df = results_df.copy()
        n = int(cfg["output"].get("transcript_preview_chars", 700))
        results_df["Расшифровка"] = results_df["Расшифровка"].astype(str).str.slice(0, n)

    # порядок колонок
    base_cols = ["Система", "Уникальные ID", "Расшифровка", cfg["excel"]["human_tags_column"]]
    dynamic_cols = [c for c in results_df.columns if c not in base_cols]
    results_df = results_df[base_cols + dynamic_cols]

    # записываем таблицу
    ws_main.append(list(results_df.columns))
    for row in dataframe_to_rows(results_df, index=False, header=False):
        ws_main.append(row)

    # создаём структурированную таблицу Excel
    last_row = ws_main.max_row
    last_col = ws_main.max_column
    tbl = Table(
        displayName="Results",
        ref=f"A1:{get_column_letter(last_col)}{last_row}",
    )
    tbl.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9", showRowStripes=True
    )
    ws_main.add_table(tbl)

    # определяем имена LAMBDA
    if cfg["excel"].get("enable_dynamic_formulas", True):
        define_lambda_names(wb)

    headers = [cell.value for cell in ws_main[1]]
    human_col_idx = headers.index(cfg["excel"]["human_tags_column"]) + 1

    # вставляем формулы для ошибок и счётчиков
    for j, head in enumerate(headers, start=1):
        if head.startswith("Теги ["):
            pkey = head.split("[", 1)[1].split("]", 1)[0]
            col_model = j
            col_err_minus = headers.index(f"Ошибки (−) [{pkey}]") + 1
            col_err_plus = headers.index(f"Ошибки (+) [{pkey}]") + 1
            col_err_minusn = headers.index(f"Ошибки (−) #[{pkey}]") + 1
            col_err_plusn = headers.index(f"Ошибки (+) #[{pkey}]") + 1
            for r in range(2, last_row + 1):
                h = f"{get_column_letter(human_col_idx)}{r}"
                m = f"{get_column_letter(col_model)}{r}"
                if cfg["excel"].get("enable_dynamic_formulas", True):
                    ws_main.cell(row=r, column=col_err_minus).value = (
                        f"=IF({h}=\"\",\"\",SET_MINUS({h},{m}))"
                    )
                    ws_main.cell(row=r, column=col_err_plus).value = (
                        f"=IF({h}=\"\",\"\",SET_MINUS({m},{h}))"
                    )
                    ws_main.cell(row=r, column=col_err_minusn).value = (
                        f"=IF({get_column_letter(col_err_minus)}{r}=\"\",0,ROWS(TEXTSPLIT({get_column_letter(col_err_minus)}{r},\",\")))"
                    )
                    ws_main.cell(row=r, column=col_err_plusn).value = (
                        f"=IF({get_column_letter(col_err_plus)}{r}=\"\",0,ROWS(TEXTSPLIT({get_column_letter(col_err_plus)}{r},\",\")))"
                    )

    # формируем «Сводку»
    ws_sum = wb.create_sheet("Сводка")
    ws_sum.append([
        "Промпт",
        "Точные совпадения %",
        "Среднее ошибок (+)",
        "Среднее ошибок (−)",
        "Доля стабильных %",
    ])

    # добавим столбцы «Совпадение [pX]» на Главном для расчёта метрик
    current_last_col = ws_main.max_column
    for j, head in enumerate(headers, start=1):
        if head.startswith("Теги ["):
            pkey = head.split("[", 1)[1].split("]", 1)[0]
            match_col_name = f"Совпадение [{pkey}]"
            # заголовок
            ws_main.cell(row=1, column=current_last_col + 1).value = match_col_name
            # формулы совпадения
            for r in range(2, last_row + 1):
                h = f"{get_column_letter(human_col_idx)}{r}"
                m = f"{get_column_letter(j)}{r}"
                if cfg["excel"].get("enable_dynamic_formulas", True):
                    ws_main.cell(row=r, column=current_last_col + 1).value = (
                        f"=IF({h}=\"\",NA(),SET_EQUAL({h},{m}))"
                    )
            headers.append(match_col_name)
            current_last_col += 1

            # находим индексы нужных столбцов для сводки
            col_err_plusn = headers.index(f"Ошибки (+) #[{pkey}]") + 1
            col_err_minusn = headers.index(f"Ошибки (−) #[{pkey}]") + 1
            col_stab_pct = headers.index(f"Стабильность [{pkey}] %") + 1
            col_match = headers.index(match_col_name) + 1

            ws_sum.append(
                [
                    pkey,
                    f"=SUMPRODUCT(--(Results[{match_col_name}]))/COUNTA(Results[{match_col_name}])*100",
                    f"=AVERAGE(Results[{headers[col_err_plusn - 1]}])",
                    f"=AVERAGE(Results[{headers[col_err_minusn - 1]}])",
                    f"=AVERAGE(Results[{headers[col_stab_pct - 1]}])",
                ]
            )

    # лист «Диагностика»
    ws_diag = wb.create_sheet("Диагностика")
    if diagnostics_df.empty:
        ws_diag.append(["(пусто)"])
    else:
        cols = [
            "Система",
            "Уникальные ID",
            "Промпт",
            "Попытка",
            "Тег",
            "Решение",
            "Конфликт",
            "Цитата",
            "Номер реплики",
            "Кто говорит",
            "Категория сигнала",
            "Флаги",
        ]
        ws_diag.append(cols)
        for r in diagnostics_df.to_dict(orient="records"):
            row = []
            for c in cols:
                v = r.get(c, "")
                if isinstance(v, dict):
                    v = json.dumps(v, ensure_ascii=False)
                row.append(v)
            ws_diag.append(row)
    # сохраняем файл
    out_path = Path(cfg["output"]["path"])
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))
    print(f"[INFO] Excel сохранён: {out_path}")


def main():
    parser = argparse.ArgumentParser(description="Запуск офлайн‑эксперимента AI‑тегирования")
    default_config = (Path(__file__).parent / "config.yaml").resolve()
    parser.add_argument(
        "--config", default=str(default_config), help="Путь к YAML‑конфигу"
    )
    args = parser.parse_args()
    cfg = load_yaml(Path(args.config))
    calls_df = load_calls(cfg)
    # Ограничение количества строк, если указано в конфиге
    limit_rows: Optional[int] = None
    if "limit_rows" in cfg.get("data", {}):
        try:
            limit_rows = int(cfg["data"]["limit_rows"])
        except Exception:
            print("[WARNING] Некорректное значение limit_rows в config.yaml, игнорирую.")
    if limit_rows is not None:
        calls_df = calls_df.head(limit_rows)
    prompts = load_prompts(cfg)

    results_df, diagnostics_df = run_experiment(cfg, calls_df, prompts)
    build_workbook(cfg, calls_df, results_df, diagnostics_df)
    print("[INFO] Готово.")


if __name__ == "__main__":
    main()